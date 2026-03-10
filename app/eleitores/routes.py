import csv
import io
from datetime import datetime
from flask import render_template, redirect, url_for, flash, request, Response, jsonify
from flask_login import login_required, current_user
from app.eleitores import bp
from app.models import Eleitor, Bairro, User, Acao, db


CLASSIFICACOES = [
    ('simpatizante', 'Simpatizante'),
    ('ativista', 'Ativista'),
    ('lideranca', 'Liderança'),
    ('cabo_eleitoral', 'Cabo Eleitoral'),
    ('candidato', 'Candidato'),
]

STATUS = [
    ('ativo', 'Ativo'),
    ('inativo', 'Inativo'),
    ('convertido', 'Convertido'),
]

FILIACOES = [
    ('nao_filiado', 'Não Filiado'),
    ('filiado', 'Filiado ao PSDB'),
]


@bp.route('/')
@login_required
def index():
    page = request.args.get('page', 1, type=int)
    q = request.args.get('q', '').strip()
    bairro_id = request.args.get('bairro', type=int)
    classificacao = request.args.get('classificacao', '')
    status = request.args.get('status', '')
    filiacao = request.args.get('filiacao', '')

    query = Eleitor.query
    if q:
        query = query.filter(Eleitor.nome.ilike(f'%{q}%'))
    if bairro_id:
        query = query.filter_by(bairro_id=bairro_id)
    if classificacao:
        query = query.filter_by(classificacao=classificacao)
    if status:
        query = query.filter_by(status=status)
    if filiacao:
        query = query.filter_by(filiacao=filiacao)

    eleitores = query.order_by(Eleitor.nome).paginate(page=page, per_page=25, error_out=False)
    bairros = Bairro.query.order_by(Bairro.nome).all()

    return render_template('eleitores/index.html',
                           eleitores=eleitores, bairros=bairros,
                           classificacoes=CLASSIFICACOES, status_list=STATUS,
                           filiacoes=FILIACOES, q=q,
                           filtro_bairro=bairro_id, filtro_class=classificacao,
                           filtro_status=status, filtro_filiacao=filiacao)


@bp.route('/novo', methods=['GET', 'POST'])
@login_required
def novo():
    bairros = Bairro.query.order_by(Bairro.nome).all()
    liderancas = Eleitor.query.filter(Eleitor.classificacao.in_(['lideranca', 'cabo_eleitoral', 'candidato'])).order_by(Eleitor.nome).all()
    if request.method == 'POST':
        eleitor = Eleitor()
        _preencher_eleitor(eleitor, request.form)
        eleitor.criado_por_id = current_user.id
        db.session.add(eleitor)
        db.session.commit()

        # Registra ação de cadastro
        acao = Acao(tipo='cadastro', descricao='Eleitor cadastrado no sistema',
                    eleitor_id=eleitor.id, responsavel_id=current_user.id)
        db.session.add(acao)
        db.session.commit()

        flash(f'Eleitor "{eleitor.nome}" cadastrado com sucesso!', 'success')
        return redirect(url_for('eleitores.detalhe', id=eleitor.id))
    return render_template('eleitores/form.html', eleitor=None,
                           bairros=bairros, liderancas=liderancas,
                           classificacoes=CLASSIFICACOES, status_list=STATUS,
                           filiacoes=FILIACOES)


@bp.route('/<int:id>')
@login_required
def detalhe(id):
    eleitor = Eleitor.query.get_or_404(id)
    acoes = Acao.query.filter_by(eleitor_id=id).order_by(Acao.data.desc()).limit(20).all()
    comunicacoes = eleitor.comunicacoes.order_by(db.desc('enviado_em')).limit(10).all()
    return render_template('eleitores/detalhe.html', eleitor=eleitor,
                           acoes=acoes, comunicacoes=comunicacoes,
                           classificacoes=CLASSIFICACOES, status_list=STATUS,
                           filiacoes=FILIACOES)


@bp.route('/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar(id):
    eleitor = Eleitor.query.get_or_404(id)
    bairros = Bairro.query.order_by(Bairro.nome).all()
    liderancas = Eleitor.query.filter(Eleitor.classificacao.in_(['lideranca', 'cabo_eleitoral', 'candidato'])).filter(Eleitor.id != id).order_by(Eleitor.nome).all()
    if request.method == 'POST':
        _preencher_eleitor(eleitor, request.form)
        eleitor.atualizado_em = datetime.utcnow()
        db.session.commit()
        flash('Dados atualizados com sucesso!', 'success')
        return redirect(url_for('eleitores.detalhe', id=eleitor.id))
    return render_template('eleitores/form.html', eleitor=eleitor,
                           bairros=bairros, liderancas=liderancas,
                           classificacoes=CLASSIFICACOES, status_list=STATUS,
                           filiacoes=FILIACOES)


@bp.route('/<int:id>/excluir', methods=['POST'])
@login_required
def excluir(id):
    if not current_user.is_coordenador():
        flash('Sem permissão para excluir eleitores.', 'danger')
        return redirect(url_for('eleitores.index'))
    eleitor = Eleitor.query.get_or_404(id)
    nome = eleitor.nome
    db.session.delete(eleitor)
    db.session.commit()
    flash(f'Eleitor "{nome}" removido.', 'warning')
    return redirect(url_for('eleitores.index'))


@bp.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    if not current_user.is_coordenador():
        flash('Sem permissão para importar dados.', 'danger')
        return redirect(url_for('eleitores.index'))
    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        if not arquivo or not arquivo.filename.endswith('.csv'):
            flash('Envie um arquivo CSV válido.', 'danger')
            return redirect(url_for('eleitores.importar'))

        stream = io.StringIO(arquivo.stream.read().decode('utf-8-sig'))
        reader = csv.DictReader(stream)
        bairros_cache = {b.nome.lower(): b for b in Bairro.query.all()}
        importados = 0
        erros = []

        for i, row in enumerate(reader, start=2):
            try:
                nome = row.get('nome', '').strip()
                if not nome:
                    erros.append(f'Linha {i}: nome vazio.')
                    continue
                bairro_nome = row.get('bairro', '').strip().lower()
                bairro = bairros_cache.get(bairro_nome)

                eleitor = Eleitor(
                    nome=nome,
                    telefone=row.get('telefone', '').strip(),
                    email=row.get('email', '').strip() or None,
                    bairro_id=bairro.id if bairro else None,
                    zona_eleitoral=row.get('zona_eleitoral', '').strip() or None,
                    secao_eleitoral=row.get('secao_eleitoral', '').strip() or None,
                    titulo_numero=row.get('titulo_numero', '').strip() or None,
                    titulo_zona=row.get('titulo_zona', '').strip() or None,
                    titulo_secao=row.get('titulo_secao', '').strip() or None,
                    classificacao=row.get('classificacao', 'simpatizante').strip(),
                    filiacao=row.get('filiacao', 'nao_filiado').strip(),
                    criado_por_id=current_user.id,
                )
                cpf_raw = row.get('cpf', '').strip()
                if cpf_raw:
                    eleitor.cpf = cpf_raw
                nasc_raw = row.get('nascimento', '').strip()
                if nasc_raw:
                    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                        try:
                            eleitor.nascimento = datetime.strptime(nasc_raw, fmt).date()
                            break
                        except ValueError:
                            continue
                db.session.add(eleitor)
                importados += 1
            except Exception as e:
                erros.append(f'Linha {i}: {str(e)}')

        db.session.commit()
        flash(f'{importados} eleitores importados com sucesso!', 'success')
        if erros:
            for e in erros[:10]:
                flash(e, 'warning')
        return redirect(url_for('eleitores.index'))

    return render_template('eleitores/importar.html')


@bp.route('/modelo-csv')
@login_required
def modelo_csv():
    headers = ['nome', 'telefone', 'email', 'nascimento', 'cpf', 'bairro',
               'zona_eleitoral', 'secao_eleitoral', 'titulo_numero', 'titulo_zona',
               'titulo_secao', 'classificacao', 'filiacao']
    exemplo = ['João da Silva', '(11) 98765-4321', 'joao@exemplo.com',
               '15/03/1980', '000.000.000-00', 'Centro',
               '001', '0001', '1234567890', '001', '0001', 'simpatizante', 'nao_filiado']
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(exemplo)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=modelo_eleitores.csv'}
    )


@bp.route('/exportar')
@login_required
def exportar():
    if not current_user.is_coordenador():
        flash('Sem permissão para exportar dados.', 'danger')
        return redirect(url_for('eleitores.index'))
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = 'Eleitores'

    headers = ['ID', 'Nome', 'Telefone', 'E-mail', 'Nascimento', 'Bairro',
               'Zona Eleitoral', 'Seção', 'Classificação', 'Status', 'Filiação',
               'Título Nº', 'Cadastrado em']
    header_fill = PatternFill(start_color='005BB5', end_color='005BB5', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for eleitor in Eleitor.query.order_by(Eleitor.nome).all():
        ws.append([
            eleitor.id, eleitor.nome, eleitor.telefone, eleitor.email,
            eleitor.nascimento.strftime('%d/%m/%Y') if eleitor.nascimento else '',
            eleitor.bairro.nome if eleitor.bairro else '',
            eleitor.zona_eleitoral, eleitor.secao_eleitoral,
            eleitor.classificacao, eleitor.status, eleitor.filiacao,
            eleitor.titulo_numero,
            eleitor.criado_em.strftime('%d/%m/%Y') if eleitor.criado_em else '',
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=eleitores_psdb.xlsx'}
    )


def _preencher_eleitor(eleitor, form):
    eleitor.nome = form.get('nome', '').strip()
    eleitor.telefone = form.get('telefone', '').strip() or None
    eleitor.email = form.get('email', '').strip() or None
    eleitor.genero = form.get('genero', '').strip() or None
    eleitor.endereco = form.get('endereco', '').strip() or None
    eleitor.observacoes = form.get('observacoes', '').strip() or None
    eleitor.classificacao = form.get('classificacao', 'simpatizante')
    eleitor.status = form.get('status', 'ativo')
    eleitor.filiacao = form.get('filiacao', 'nao_filiado')

    bairro_id = form.get('bairro_id', type=int) if hasattr(form, 'get') else None
    if bairro_id:
        eleitor.bairro_id = bairro_id

    eleitor.zona_eleitoral = form.get('zona_eleitoral', '').strip() or None
    eleitor.secao_eleitoral = form.get('secao_eleitoral', '').strip() or None
    eleitor.titulo_numero = form.get('titulo_numero', '').strip() or None
    eleitor.titulo_zona = form.get('titulo_zona', '').strip() or None
    eleitor.titulo_secao = form.get('titulo_secao', '').strip() or None

    ponto_focal_id = form.get('ponto_focal_id', type=int) if hasattr(form, 'get') else None
    eleitor.ponto_focal_id = ponto_focal_id or None

    cpf_raw = form.get('cpf', '').strip()
    if cpf_raw:
        eleitor.cpf = cpf_raw

    nasc_raw = form.get('nascimento', '').strip()
    if nasc_raw:
        try:
            eleitor.nascimento = datetime.strptime(nasc_raw, '%Y-%m-%d').date()
        except ValueError:
            pass
